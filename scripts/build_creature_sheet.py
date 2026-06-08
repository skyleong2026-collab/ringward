import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
DARK = "1A1420"; HEAD = "2C2430"; STRIPE = "F4F1F6"; INK = "201A26"
thin = Side(style="thin", color="D8D2DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# (name, type, role, found_in, color_hex, look_hook, mj_subject)
TYPE_COLOR = {
    "Reactor": "FF5A2A", "Bulwark": "7FD6FF", "Mender": "7ED321", "Booster": "B06BFF",
    "Striker": "FFD166", "Assassin": "FF7A9C", "Warden": "8FD8FF", "Hexer": "9B59D0",
}
ROLE = {
    "Reactor": "Powers up, then blows up", "Bulwark": "Soaks hits, guards the team",
    "Mender": "Keeps the squad alive", "Booster": "Makes an ally hit way harder",
    "Striker": "Fast — lots of quick hits", "Assassin": "Hunts and finishes the weak",
    "Warden": "Freezes enemies out of their turns", "Hexer": "Makes enemies take more from everyone",
}
CREATURES = [
    ("Fizzpop","Reactor","The Crackling Outer Ring",
     "A jittery firecracker that can't sit still — hoards fire, then lets it all go at once. Sparks leak from cracked-stone seams; an over-bright core flares.",
     "a small jittery ember grunling, restless wide eyes, sparks leaking from cracked-stone seams, an over-bright molten core, a fuse-like tail, looks about to burst"),
    ("Glowtail","Reactor","Dusk-lit hollows near the edge",
     "Slow-burning and patient — its long ember-lantern tail pulses brighter the longer a fight drags on. Calmer menace than Fizzpop.",
     "a smouldering grunling with a long glowing ember-lantern tail that brightens over time, deep warm core, patient and slow-burning, embers trailing behind it"),
    ("Cinderpaw","Reactor","Old burn-scars on the approach",
     "An ember-cat that walks on coals — whatever it touches smoulders for days. Ash-grey cracked-stone 'fur' with molten seams, smoking paws.",
     "a small ember-cat grunling, ash-grey cracked-stone fur with molten glowing seams, paws that smoulder and leave scorch-prints, a bright fire core in its chest"),
    ("Stoneward","Bulwark","The rubble of a fallen gate",
     "A walking wall — siege engines broke against it and it never moved. Massive, squat, mossy granite slabs; a steady cool-blue core deep inside.",
     "a massive squat slab-grunling, a walking wall of mossy granite plates, heavy and immovable, a steady cool-blue core deep in its chest, ancient and patient"),
    ("Ironwall","Bulwark","A sealed keep, deeper in",
     "Older than the blight. It guards, it does not attack — and it does not fall. Ancient riveted iron-and-stone sentinel, gate-like and monumental.",
     "an ancient sentinel grunling of riveted iron and weathered stone, gate-like and monumental, a dim old core, built to guard and never fall"),
    ("Mossback","Mender","A green seam between the rings",
     "Where it sleeps, the wounded wake whole — hunters leave it be. A gentle walking garden: soft green growth blanketing old stone, a calm restorative glow.",
     "a gentle mossy grunling, soft green growth blanketing old stone, a calm restorative green glow in its chest, sleepy and kind, a small walking garden"),
    ("Dewleaf","Mender","Mist-fed glades at first light",
     "Weeps a water that knits flesh — rare, and quick to flee. Delicate leaf-and-dew creature, translucent and timid, dewdrops falling, poised to bolt.",
     "a delicate timid grunling of leaf and dew, translucent water-bright features, a soft green-blue glow, dewdrops falling from it, poised to flee"),
    ("Buzzline","Booster","Storm-wire thickets",
     "Latches onto a stronger creature and makes its blows land twice as hard. Small and wiry, wound in crackling storm-wire, clamp-limbs to grab on.",
     "a small wiry grunling wound in crackling storm-wire, electric filaments arcing, clamp-like limbs to latch onto allies, an excitable bright violet core"),
    ("Tanglewing","Booster","Wind-tunnels between cliffs",
     "Hums a note that lifts a whole pack at once — never fights alone. A conductor: vine-and-wire wings, visible sound-ripples, a warm resonant glow.",
     "a winged grunling that hums an uplifting note, visible sound-ripples in the air around it, vine-and-wire wings, a warm resonant violet core, conductor of the pack"),
    ("Swiftpaw","Striker","The fast, narrow trails",
     "Gone before you see it move — a dozen strikes in a breath. Lean, sleek, aerodynamic stone with gold accents, coiled to spring, faint motion-streaks.",
     "a lean fast quadruped grunling, sleek aerodynamic stone body with gold accents, coiled and quick, faint motion streaks, built for a dozen strikes in a breath"),
    ("Dartwing","Striker","High ledges on the inner climb",
     "Always acts first — the fastest thing anyone's lived to describe. An arrow-shaped winged darter, gold-tipped, razor-light, hyper-alert.",
     "a darting arrow-shaped winged grunling, gold-tipped and razor-light, hyper-alert eyes, built to always strike first, a blur of speed"),
    ("Shadefang","Assassin","The deep dark past the Warden",
     "Hunts only the wounded; waits for the kill, then takes it. Low, smoke-wreathed dark-stone predator, a single cold rose-pink core, long fangs.",
     "a low shadowy predator grunling, smoke-wreathed dark stone body, a cold rose-pink glow at its core, long fangs, patient and lethal"),
    ("Veilclaw","Assassin","The lightless inner rings",
     "A killer wrapped in shadow — the weaker you are, the harder it bites. Half-dissolved into a dark veil, hooked claws, a faint crimson core.",
     "a grunling half-dissolved into shadow, a veil of darkness over hooked claws, a faint crimson core, near-invisible and terrifying"),
    ("Frostward","Warden","The frostbound deep, far inward",
     "A cold so total it stops a creature mid-step; time seems to freeze around it. Frost-armoured rime-coated stone, cold mist pouring off, a pale-blue still core.",
     "a frost-armoured grunling of rime-coated stone, cold mist pouring off it, a pale-blue still core, ice creeping from its feet, a presence that freezes time"),
    ("Rimecaller","Warden","The silent rings near the Drop",
     "Calls a stillness down over a whole battlefield — few have seen one and moved again. Taller, ceremonial, an ice-crystal crown, a serene glacial glow.",
     "a tall ceremonial frost grunling with an ice-crystal crown, glacial-blue glow, conductor of a battlefield stillness, serene and dreadful, snow stilled in the air"),
    ("Blightcap","Hexer","The witherfen, where the blight pools",
     "Where it treads, armour rots and wounds refuse to close — the blight seems to wear its face. A fungal cap over corroded stone, sickly violet spores.",
     "a fungal blight-grunling, a diseased mushroom cap over corroded stone, sickly violet spores drifting off it, a corrupted glowing core, rot follows its steps"),
    ("Hexmoth","Hexer","The spore-choked dark",
     "Its dust settles on you like a mark — everything that strikes you after bites deeper. A moth-grunling, dusty violet wings, glowing eye-spots.",
     "a moth-like grunling with dusty violet wings shedding cursed powder, glowing eye-spots on its wings, an ominous soft-purple core, it marks all it touches"),
]

STYLE_SUFFIX = ("full-body character concept, single creature centered, plain neutral background, "
    "even soft lighting, clear readable silhouette, muted folk-mystery storybook style, hand-painted "
    "painterly, warm-but-melancholy palette, an earthen-machine creature with a glowing core-heart in "
    "its chest, consistent friendly proportions, no text, no UI --ar 1:1")

wb = Workbook()

# ── Sheet 1: printable reference ──
s = wb.active; s.title = "Creatures"
s.sheet_view.showGridLines = False
title = "RINGWARD  —  Creature Design Reference"
s["A1"] = title
s["A1"].font = Font(name=FONT, size=16, bold=True, color="FFFFFF")
s["A1"].fill = PatternFill("solid", fgColor=DARK)
s.merge_cells("A1:F1")
s.row_dimensions[1].height = 26
s["A2"] = ("17 grunlings — earthen-machine creatures with a glowing core-heart. This is your springboard, "
           "not a verdict: lock each look in Midjourney, then feed that single image to PixelLab. Edit freely.")
s["A2"].font = Font(name=FONT, size=9, italic=True, color="6A6076")
s.merge_cells("A2:F2"); s.row_dimensions[2].height = 28
s["A2"].alignment = Alignment(wrap_text=True, vertical="center")

headers = ["#", "Name", "Type / Role", "Found in (ring)", "Colour", "Look & personality  (your starting point)"]
hr = 4
for c, h in enumerate(headers, 1):
    cell = s.cell(row=hr, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEAD)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border
s.row_dimensions[hr].height = 20

for i, (name, typ, found, hook, _mj) in enumerate(CREATURES):
    r = hr + 1 + i
    stripe = STRIPE if i % 2 else "FFFFFF"
    row = [i + 1, name, f"{typ} · {ROLE[typ]}", found, f"#{TYPE_COLOR[typ]}", hook]
    for c, val in enumerate(row, 1):
        cell = s.cell(row=r, column=c, value=val)
        cell.font = Font(name=FONT, size=9.5, color=INK, bold=(c == 2))
        cell.fill = PatternFill("solid", fgColor=stripe)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 4, 6)),
                                   horizontal="center" if c in (1,) else "left")
    s.cell(row=r, column=5).fill = PatternFill("solid", fgColor=TYPE_COLOR[typ])
    s.cell(row=r, column=5).font = Font(name=FONT, size=8, color="201A26", bold=True)
    s.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
    s.row_dimensions[r].height = 52

widths = [4, 13, 22, 24, 9, 64]
for c, w in enumerate(widths, 1):
    s.column_dimensions[get_column_letter(c)].width = w
s.freeze_panes = "A5"
s.page_setup.orientation = "landscape"
s.page_setup.fitToWidth = 1; s.page_setup.fitToHeight = 0
s.sheet_properties.pageSetUpPr.fitToPage = True
s.print_title_rows = "4:4"
s.print_options.horizontalCentered = True
s.page_margins.left = s.page_margins.right = 0.3
s.page_margins.top = s.page_margins.bottom = 0.4

# ── Sheet 2: ready-to-paste Midjourney prompts ──
p = wb.create_sheet("Midjourney Prompts")
p.sheet_view.showGridLines = False
p["A1"] = "Midjourney prompts  —  one locked image per creature"
p["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
p["A1"].fill = PatternFill("solid", fgColor=DARK)
p.merge_cells("A1:C1"); p.row_dimensions[1].height = 24
notes = ("THE ONE TRICK: generate until you love a STYLE, grab its  --sref <code>  (or a reference image), "
         "and add that same --sref to EVERY prompt below — that's what makes all 17 read as one set. "
         "Aim for clean, single-character, neutral-background concepts (they feed PixelLab best). Lock one "
         "image per creature; save its prompt + seed + sref. That locked image is what you hand PixelLab.")
p["A2"] = notes
p["A2"].font = Font(name=FONT, size=9, italic=True, color="6A6076")
p.merge_cells("A2:C2"); p.row_dimensions[2].height = 56
p["A2"].alignment = Alignment(wrap_text=True, vertical="center")

phr = 4
for c, h in enumerate(["Name", "Type", "Prompt  (paste, then append your --sref)"], 1):
    cell = p.cell(row=phr, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEAD); cell.border = border
    cell.alignment = Alignment(vertical="center")
for i, (name, typ, _found, _hook, mj) in enumerate(CREATURES):
    r = phr + 1 + i
    stripe = STRIPE if i % 2 else "FFFFFF"
    full = f"{mj}, {STYLE_SUFFIX}"
    for c, val in enumerate([name, typ, full], 1):
        cell = p.cell(row=r, column=c, value=val)
        cell.font = Font(name=FONT, size=9 if c == 3 else 9.5, color=INK, bold=(c == 1))
        cell.fill = PatternFill("solid", fgColor=stripe); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
    p.row_dimensions[r].height = 78
for c, w in enumerate([13, 11, 100], 1):
    p.column_dimensions[get_column_letter(c)].width = w
p.freeze_panes = "A5"
p.page_setup.orientation = "landscape"
p.sheet_properties.pageSetUpPr = s.sheet_properties.pageSetUpPr

out_dir = os.path.expanduser("~/Desktop")
if not os.path.isdir(out_dir):
    out_dir = os.path.expanduser("~")
out = os.path.join(out_dir, "Ringward-Creature-Reference.xlsx")
wb.save(out)
print("SAVED:", out)
